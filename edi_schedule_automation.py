import os
import glob
import csv
from datetime import datetime
import openpyxl
from openpyxl import Workbook

# Developed by Eva Karakostas using Python.

# ---------------------------------
# INPUT & OUTPUT FOLDERS
# ---------------------------------

BASE_INPUT = r"C:\Demo\Customer\Order Schedules"
BASE_OUTPUT = r"C:\Demo\Customer\EDI Templates"

# Automatically detect latest year folder
year_folders = [
    f for f in os.listdir(BASE_INPUT)
    if f.isdigit() and os.path.isdir(os.path.join(BASE_INPUT, f))
]

if not year_folders:
    raise FileNotFoundError("No year folders found.")

latest_year = max(year_folders)

FOLDER = os.path.join(BASE_INPUT, latest_year)
EDI_BASE_FOLDER = os.path.join(BASE_OUTPUT, latest_year)

print(f"Using Year Folder: {latest_year}")

# ---------------------------------
# Find newest non-temp Excel File
# ---------------------------------

xlsx_files = [
    f for f in glob.glob(os.path.join(FOLDER, "*.xlsx"))
    if not os.path.basename(f).startswith("~$")
]

if not xlsx_files:
    raise FileNotFoundError("No Excel files found.")

excel_file = max(xlsx_files, key=os.path.getmtime)
print(f"Using Excel file: {excel_file}")

# ---------------------------------
# Load Excel
# ---------------------------------

wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb.active

# ---------------------------------
# Auto-detect Header Row
# ---------------------------------

header_row = None
headers = None

for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row:
        row_values = [str(c).strip() if c else "" for c in row]
        if "Part Type" in row_values and "Quantity" in row_values:
            header_row = i
            headers = row_values
            break

if not header_row:
    raise ValueError("Header row not found.")

col = {name: idx for idx, name in enumerate(headers)}

# ---------------------------------
# Part Configuration
# ---------------------------------

PARTS = {
    "PART_A": {
        "filename": "Customer862.PartA",
        "lin": "LIN**BP*PARTA0001***PO*PARTAPO*PD*PART A DESCRIPTION*PL*1~"
    },
    "PART_B": {
        "filename": "Customer862.PartB",
        "lin": "LIN**BP*PARTB0001***PO*PARTBPO*PD*PART B DESCRIPTION*PL*1~"
    },
    "PART_C": {
        "filename": "Customer862.PartC",
        "lin": "LIN**BP*PARTC0001***PO*PARTCPO*PD*PART C DESCRIPTION*PL*1~"
    },
    "PART_D": {
        "filename": "Customer862.PartD",
        "lin": "LIN**BP*PARTD0001***PO*PARTDPO*PD*PART D DESCRIPTION*PL*1~"
    }
}

# ---------------------------------
# Helpers
# ---------------------------------

def fmt(d):
    return d.strftime("%Y%m%d") if isinstance(d, datetime) else None

def fmt_folder_date(d):
    """Convert 'YYYYMMDD' string to 'YYYY.MM.DD' for folder name"""
    return f"{d[:4]}.{d[4:6]}.{d[6:]}" if d else ""

today = datetime.today()
today_edi = today.strftime("%y%m%d")
today_long = today.strftime("%Y%m%d")
today_file = today.strftime("%Y.%m.%d")

# ---------------------------------
# Determine Delivery Range for Folder Naming
# ---------------------------------

all_delivers = []
for part, cfg in PARTS.items():
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if r[col["Part Type"]] == part:
            qty = r[col["Quantity"]]
            if qty and qty != 0:
                all_delivers.append(fmt(r[col["Deliver Date"]]))

if all_delivers:
    dl_start = min(all_delivers)
    dl_end = max(all_delivers)
else:
    dl_start = dl_end = today_long

# ---------------------------------
# Create Single Weekly EDI Output Folder
# ---------------------------------

week_folder_name = f"{today_file} ({fmt_folder_date(dl_start)} ~ {fmt_folder_date(dl_end)} Orders)"
week_folder_path = os.path.join(EDI_BASE_FOLDER, week_folder_name)
os.makedirs(week_folder_path, exist_ok=True)

# ---------------------------------
# Process each part separately
# ---------------------------------

for part, cfg in PARTS.items():
    rows = []

    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if r[col["Part Type"]] != part:
            continue

        qty = r[col["Quantity"]]
        if not qty or qty == 0:
            continue

        rows.append({
            "qty": int(qty),
            "po": r[col["PO#"]],
            "deliver": fmt(r[col["Deliver Date"]])
        })

    if not rows:
        print(f"Skipping {part} (no shipments)")
        continue

    dl_start_part = min(r["deliver"] for r in rows)
    dl_end_part = max(r["deliver"] for r in rows)

    edi = [
        f"ISA*00*          *00*          *ZZ*SENDERID       *ZZ*RECEIVERID     *{today_edi}*2121*U*00401*000000059*0*P*>~",
        f"GS*SS*SENDERID*RECEIVERID*{today_long}*2121*59*X*004010~",
        "ST*862*000059001~",
        f"BSS*05*{today_long}000059001*{today_long}*DL*{dl_start_part}*{dl_end_part}~",
        f"DTM*102*{today_long}*2121~",
        "N1*SU*SUPPLIER NAME*92*SUPPLIERID~",
        "N1*ST*SHIP-TO CUSTOMER NAME*92*CUSTOMERID~",
        "REF*DK*CUSTOMERID~",
        cfg["lin"],
        "UIT*PC~"
    ]

    for r in rows:
        edi.extend([
            f"FST*{r['qty']}*C*C*{r['deliver']}~",
            f"JIT*{r['qty']}*000001~" if part == "PART_D" else f"JIT*{r['qty']}~",
            f"REF*RE*{r['po']}~"
])

    if part in ("PART_D", "PART_A", "PART_B"):
        edi.extend([
            "SHP*01*00000*050*YYYYMMDD~",
            "REF*SI*REFERENCEID~",
            "SHP*02*00000*051*YYYYMMDD**YYYYMMDD~"
        ])

    # ---------------------------------
    # Calculate proper SE segment
    # ---------------------------------
    
    edi.append("CTT*1*00000~")

    st_index = next(i for i, s in enumerate(edi) if s.startswith("ST"))
    se_count = len(edi) - st_index + 1

    edi.append(f"SE*{se_count}*000059001~")
    edi.append("GE*1*59~")
    edi.append("IEA*1*000000059~")

    # ---------------------------------
    # Save as CSV (ERP readable)
    # ---------------------------------
    
    out_name = f"{cfg['filename']}.{today_file}.csv"
    out_path = os.path.join(week_folder_path, out_name)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for line in edi:
            writer.writerow([line])

    print(f"Created: {out_path}")
