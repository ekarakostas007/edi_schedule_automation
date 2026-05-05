import os
import glob
import csv
from datetime import datetime
import openpyxl
from openpyxl import Workbook

# Developed by Eva Karakostas using Python.

# ---------------------------------
# INPUT & OUTPUT FOLDERS
# ---------------------------------

BASE_INPUT = r"S:\Sales\Customer\Toyota\Shipping\TMMTN Truck Schedules_Liners"
BASE_OUTPUT = r"S:\Sales\Customer\Toyota\Shipping\Plex EDI Templates"

# Automatically detect latest year folder
year_folders = [
    f for f in os.listdir(BASE_INPUT)
    if f.isdigit() and os.path.isdir(os.path.join(BASE_INPUT, f))
]

if not year_folders:
    raise FileNotFoundError("No year folders found.")

latest_year = max(year_folders)

FOLDER = os.path.join(BASE_INPUT, latest_year)
EDI_BASE_FOLDER = os.path.join(BASE_OUTPUT, latest_year)

print(f"Using Year Folder: {latest_year}")

# ---------------------------------
# Find newest non-temp Excel File
# ---------------------------------

xlsx_files = [
    f for f in glob.glob(os.path.join(FOLDER, "*.xlsx"))
    if not os.path.basename(f).startswith("~$")
]

if not xlsx_files:
    raise FileNotFoundError("No Excel files found.")

excel_file = max(xlsx_files, key=os.path.getmtime)
print(f"Using Excel file: {excel_file}")

# ---------------------------------
# Load Excel
# ---------------------------------

wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb.active

# ---------------------------------
# Auto-detect Header Row
# ---------------------------------

header_row = None
headers = None

for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row:
        row_values = [str(c).strip() if c else "" for c in row]
        if "Liner Type" in row_values and "Quantity" in row_values:
            header_row = i
            headers = row_values
            break

if not header_row:
    raise ValueError("Header row not found.")

col = {name: idx for idx, name in enumerate(headers)}

# ---------------------------------
# Part Configuration
# ---------------------------------

PARTS = {
    "2.4L": {
        "filename": "Bodine862.L4T-115E",
        "lin": "LIN**BP*11461F0030***PO*115EPO*PD*115E 2.4LT Liner*PL*1~"
    },
    "V6-T": {
        "filename": "Bodine862.V6TThinnerLiner",
        "lin": "LIN**BP*11461F4011***PO*103EPO*PD*103E/978F*PL*1~"
    },
    "M20A": {
        "filename": "Bodine862.M20AB",
        "lin": "LIN**BP*11461F2010-B***PO*M20APO*PD*M20A/984F*PL*1~"
    },
    "TNGA": {
        "filename": "Bodine862.A25A",
        "lin": "LIN**BP*11461F0010***PO*A25APO*PD*A25A/946F*PL*1~"
    }
}

# ---------------------------------
# Helpers
# ---------------------------------

def fmt(d):
    return d.strftime("%Y%m%d") if isinstance(d, datetime) else None

def fmt_folder_date(d):
    """Convert 'YYYYMMDD' string to 'YYYY.MM.DD' for folder name"""
    return f"{d[:4]}.{d[4:6]}.{d[6:]}" if d else ""

today = datetime.today()
today_edi = today.strftime("%y%m%d")
today_long = today.strftime("%Y%m%d")
today_file = today.strftime("%Y.%m.%d")

# ---------------------------------
# Determine Delivery Range for Folder Naming
# ---------------------------------

all_delivers = []
for part, cfg in PARTS.items():
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if r[col["Liner Type"]] == part:
            qty = r[col["Quantity"]]
            if qty and qty != 0:
                all_delivers.append(fmt(r[col["Deliver Date"]]))

if all_delivers:
    dl_start = min(all_delivers)
    dl_end = max(all_delivers)
else:
    dl_start = dl_end = today_long

# ---------------------------------
# Create Single Weekly EDI Output Folder
# ---------------------------------

week_folder_name = f"{today_file} ({fmt_folder_date(dl_start)} ~ {fmt_folder_date(dl_end)} Orders)"
week_folder_path = os.path.join(EDI_BASE_FOLDER, week_folder_name)
os.makedirs(week_folder_path, exist_ok=True)

# ---------------------------------
# Process each part separately
# ---------------------------------

for part, cfg in PARTS.items():
    rows = []

    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if r[col["Liner Type"]] != part:
            continue

        qty = r[col["Quantity"]]
        if not qty or qty == 0:
            continue

        rows.append({
            "qty": int(qty),
            "po": r[col["PO#"]],
            "deliver": fmt(r[col["Deliver Date"]])
        })

    if not rows:
        print(f"Skipping {part} (no shipments)")
        continue

    dl_start_part = min(r["deliver"] for r in rows)
    dl_end_part = max(r["deliver"] for r in rows)

    edi = [
        f"ISA*00*          *00*          *ZZ*7312655500     *ZZ*8474465336     *{today_edi}*2121*U*00401*000000059*0*P*>~",
        f"GS*SS*7312655500*8474465336*{today_long}*2121*59*X*004010~",
        "ST*862*000059001~",
        f"BSS*05*{today_long}000059001*{today_long}*DL*{dl_start_part}*{dl_end_part}~",
        f"DTM*102*{today_long}*2121~",
        "N1*SU*TPR FEDERAL-MOGUL TENNESSEE*92*TFT~",
        "N1*ST*BODINE ALUMINUM TENNESSEE*92*7312655500~",
        "REF*DK*7312655500~",
        cfg["lin"],
        "UIT*PC~"
    ]

    for r in rows:
        edi.extend([
            f"FST*{r['qty']}*C*C*{r['deliver']}~",
            f"JIT*{r['qty']}*000001~" if part == "TNGA" else f"JIT*{r['qty']}~",
            f"REF*RE*{r['po']}~"
])

    if part in ("TNGA", "2.4L", "V6-T"):
        edi.extend([
            "SHP*01*14400*050*20190913~",
            "REF*SI*1234567890~",
            "SHP*02*14400*051*20190913**20190913~"
        ])

    # ---------------------------------
    # Calculate proper SE segment
    # ---------------------------------
    edi.append("CTT*1*63900~")

    st_index = next(i for i, s in enumerate(edi) if s.startswith("ST"))
    se_count = len(edi) - st_index + 1

    edi.append(f"SE*{se_count}*000059001~")
    edi.append("GE*1*59~")
    edi.append("IEA*1*000000059~")

    # ---------------------------------
    # Save as CSV (ERP readable)
    # ---------------------------------
    out_name = f"{cfg['filename']}.{today_file}.csv"
    out_path = os.path.join(week_folder_path, out_name)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for line in edi:
            writer.writerow([line])

    print(f"Created: {out_path}")


